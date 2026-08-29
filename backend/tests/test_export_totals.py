from io import BytesIO

from openpyxl import load_workbook

from app.services.export_service import detailed_rows, generate_excel


def _receipts():
    return [
        {
            "id": "receipt-1",
            "receiptDate": "08/01/2026",
            "supplier": "Acme",
            "category": "Food",
            "invoiceNumber": "INV-1",
            "totalAmount": "100.00",
            "items": [
                {"name": "One", "quantity": 1, "price": 40, "tax": 6.4},
                {"name": "Two", "quantity": 1, "price": 50, "tax": 8},
            ],
        },
        {
            "id": "receipt-2",
            "receiptDate": "08/02/2026",
            "supplier": "Beta",
            "category": "Office",
            "invoiceNumber": "INV-2",
            "totalAmount": 25,
            "items": [{"name": "Three", "quantity": 1, "price": 20, "tax": 3.2}],
        },
    ]


def test_detailed_rows_do_not_repeat_receipt_total():
    rows = detailed_rows(_receipts())

    assert [row["Receipt Total"] for row in rows] == [100.0, "", 25.0]
    assert sum(row["Receipt Total"] for row in rows if row["Receipt Total"] != "") == 125.0
    assert rows[0]["Calculated Item Total"] != rows[0]["Receipt Total"]


def test_detailed_excel_has_authoritative_receipt_totals_sheet():
    workbook_bytes = generate_excel(_receipts(), "detailed")
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)

    assert "Receipt Totals" in workbook.sheetnames
    sheet = workbook["Receipt Totals"]
    assert [row[5].value for row in sheet.iter_rows(min_row=2, max_col=6)] == [100.0, 25.0]

    detail = workbook["Detailed"]
    headers = [cell.value for cell in detail[1]]
    assert "Calculated Item Total" in headers
    assert "Receipt Total" in headers
