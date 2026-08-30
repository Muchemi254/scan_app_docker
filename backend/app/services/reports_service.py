"""Comprehensive reporting engine.

Every entity in the system has a report definition (report definitions +
parameterized queries + column allowlists + format renderers).  Reports are
either owner-scoped (the requesting user sees their own data — enforced
explicitly in SQL because the app connects as a superuser that bypasses RLS)
or admin-only.  Sensitive staff columns (emails, names, uids, credentials)
are *excluded* from exports by default; only admins may opt in with
include_sensitive=true.  password_hash / secrets are never selectable.
"""

import io
import csv
import json
import logging
import re
from datetime import datetime, date, timezone
from decimal import Decimal
from typing import Any, Dict, List, Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph

from app.core.database import get_pool
from app.schemas.receipt import AuditAction, AuditFieldChange
from app.services.audit_service import AuditService

logger = logging.getLogger(__name__)

MAX_ROWS = 50_000


class ReportColumn:
    def __init__(
        self,
        col: str,
        label: str,
        sensitive: bool = False,
        money: bool = False,
    ):
        self.col = col
        self.label = label
        self.sensitive = sensitive
        self.money = money


class ReportDef:
    def __init__(
        self,
        key: str,
        name: str,
        description: str,
        scope: str,
        columns: List[ReportColumn],
        base: str,
        tenant: Optional[str] = None,
        date_col: Optional[str] = None,
        filters: Optional[Dict[str, str]] = None,
        order_by: str = "",
        group_by: Optional[str] = None,
        money_cols: Optional[List[str]] = None,
        expense_cond: Optional[str] = None,
    ):
        self.key = key
        self.name = name
        self.description = description
        self.scope = scope  # "owner" (any authenticated user) or "admin"
        self.columns = columns
        self.base = base
        self.tenant = tenant  # SQL fragment scoping rows to the request user
        self.date_col = date_col  # SQL fragment used for date_from/date_to
        self.filters = filters or {}
        self.order_by = order_by
        self.group_by = group_by
        self.money_cols = money_cols or []
        # SQL fragment restricting to spend rows (excludes quotations,
        # proformas, deposits, notes). Set for receipt money reports.
        self.expense_cond = expense_cond


REPORT_DEFS: Dict[str, ReportDef] = {}


def _register(d: ReportDef) -> None:
    REPORT_DEFS[d.key] = d


_register(ReportDef(
    key="receipts_register",
    name="Receipt register",
    description="Every receipt with status, supplier, category, tax and document identifiers.",
    scope="owner",
    columns=[
        ReportColumn("r.id", "Receipt ID"),
        ReportColumn("r.user_id", "Owner user ID", sensitive=True),
        ReportColumn("r.status", "Status"),
        ReportColumn("r.entry_type", "Entry type"),
        ReportColumn("r.supplier", "Supplier"),
        ReportColumn("r.category", "Category"),
        ReportColumn("r.location", "Location"),
        ReportColumn("r.receipt_date", "Receipt date"),
        ReportColumn("r.total_amount", "Total amount (KES)", money=True),
        ReportColumn("r.tax_amount", "Tax amount (KES)", money=True),
        ReportColumn("r.tax_rate", "Tax rate"),
        ReportColumn("r.invoice_number", "Invoice number"),
        ReportColumn("r.kra_pin", "Supplier KRA PIN", sensitive=True),
        ReportColumn("r.buyer_kra_pin", "Buyer KRA PIN", sensitive=True),
        ReportColumn("r.cu_invoice", "CU invoice"),
        ReportColumn("r.batch_title", "Batch title"),
        ReportColumn("r.image_filename", "Image file"),
        ReportColumn("r.created_at", "Created"),
        ReportColumn("r.scanned_at", "Scanned"),
    ],
    base="receipts r",
    tenant="r.user_id",
    date_col="r.receipt_date",
    filters={
        "status": "r.status",
        "category": "r.category",
        "supplier": "r.supplier",
        "location": "r.location",
        "batch_title": "r.batch_title",
    },
    order_by=" ORDER BY r.receipt_date DESC, r.created_at DESC",
    money_cols=["Total amount (KES)", "Tax amount (KES)"],
))

_register(ReportDef(
    key="line_items",
    name="Line items",
    description="Line-item level detail across receipts: quantities, prices, tax and discounts.",
    scope="owner",
    columns=[
        ReportColumn("l.id", "Item ID"),
        ReportColumn("r.id", "Receipt ID"),
        ReportColumn("r.user_id", "Owner user ID", sensitive=True),
        ReportColumn("r.supplier", "Supplier"),
        ReportColumn("r.category", "Category"),
        ReportColumn("r.receipt_date", "Receipt date"),
        ReportColumn("l.sort_order", "Sort order"),
        ReportColumn("l.name", "Item name"),
        ReportColumn("l.quantity", "Quantity"),
        ReportColumn("l.price", "Price (KES)", money=True),
        ReportColumn("l.tax", "Tax (KES)", money=True),
        ReportColumn("l.tax_rate", "Tax rate"),
        ReportColumn("l.is_zero_rated", "Zero rated"),
        ReportColumn("l.discount", "Discount (KES)", money=True),
    ],
    base="line_items l JOIN receipts r ON r.id = l.receipt_id",
    tenant="r.user_id",
    date_col="r.receipt_date",
    filters={
        "status": "r.status",
        "category": "r.category",
    },
    order_by=" ORDER BY r.receipt_date DESC, r.id, l.sort_order",
    money_cols=["Price (KES)", "Tax (KES)", "Discount (KES)"],
    expense_cond="r.entry_type = 'expense'",
))

_register(ReportDef(
    key="tax_summary",
    name="Tax / VAT analysis",
    description="Aggregated tax posture: grouped by zero-rating and tax rate with gross, tax and discount totals.",
    scope="owner",
    columns=[
        ReportColumn("l.is_zero_rated", "Zero rated"),
        ReportColumn("l.tax_rate", "Tax rate"),
        ReportColumn("COUNT(*)", "Item count"),
        ReportColumn("SUM(l.price)", "Gross amount (KES)", money=True),
        ReportColumn("SUM(COALESCE(l.tax, 0))", "Tax amount (KES)", money=True),
        ReportColumn("SUM(COALESCE(l.discount, 0))", "Discount (KES)", money=True),
    ],
    base="line_items l JOIN receipts r ON r.id = l.receipt_id",
    tenant="r.user_id",
    date_col="r.receipt_date",
    group_by="l.is_zero_rated, l.tax_rate",
    order_by=" ORDER BY l.tax_rate, l.is_zero_rated",
    money_cols=["Gross amount (KES)", "Tax amount (KES)", "Discount (KES)"],
    expense_cond="r.entry_type = 'expense'",
))

_register(ReportDef(
    key="receipt_totals",
    name="Spend totals",
    description="Monthly and per-status totals: receipt counts, item counts and sums by period.",
    scope="owner",
    columns=[
        ReportColumn("to_char(r.receipt_date, 'YYYY-MM')", "Month"),
        ReportColumn("r.status", "Status"),
        ReportColumn("COUNT(*)", "Receipt count"),
        ReportColumn("SUM(r.total_amount)", "Total amount (KES)", money=True),
        ReportColumn("SUM(COALESCE(r.tax_amount, 0))", "Tax amount (KES)", money=True),
    ],
    base="receipts r",
    tenant="r.user_id",
    date_col="r.receipt_date",
    filters={"status": "r.status"},
    group_by="to_char(r.receipt_date, 'YYYY-MM'), r.status",
    order_by=" ORDER BY 1 DESC, r.status",
    money_cols=["Total amount (KES)", "Tax amount (KES)"],
    expense_cond="r.entry_type = 'expense'",
))

_register(ReportDef(
    key="scan_sessions",
    name="Scan sessions",
    description="Batch scan session log: images, groups, chunk plan and timestamps.",
    scope="owner",
    columns=[
        ReportColumn("s.id", "Session ID"),
        ReportColumn("s.user_id", "Owner user ID", sensitive=True),
        ReportColumn("s.title", "Title"),
        ReportColumn("s.image_count", "Image count"),
        ReportColumn("s.group_count", "Group count"),
        ReportColumn("s.chunks", "Chunk plan"),
        ReportColumn("s.created_at", "Created"),
        ReportColumn("s.updated_at", "Updated"),
    ],
    base="scan_sessions s",
    tenant="s.user_id",
    date_col="s.created_at",
    order_by=" ORDER BY s.created_at DESC",
))

_register(ReportDef(
    key="scan_session_items",
    name="Scan session items",
    description="Per-image processing detail: stage, errors, extracted receipt references.",
    scope="owner",
    columns=[
        ReportColumn("i.id", "Item ID"),
        ReportColumn("s.id", "Session ID"),
        ReportColumn("s.title", "Session title"),
        ReportColumn("s.user_id", "Owner user ID", sensitive=True),
        ReportColumn("i.item_index", "Item index"),
        ReportColumn("i.group_index", "Group index"),
        ReportColumn("i.chunk_index", "Chunk index"),
        ReportColumn("i.orig_filename", "Original file"),
        ReportColumn("i.mime", "MIME type"),
        ReportColumn("i.stage", "Stage"),
        ReportColumn("i.message", "Message"),
        ReportColumn("i.error_code", "Error code"),
        ReportColumn("i.error_message", "Error message"),
        ReportColumn("i.receipt_id", "Receipt ID"),
    ],
    base="scan_session_items i JOIN scan_sessions s ON s.id = i.session_id",
    tenant="s.user_id",
    date_col="s.created_at",
    order_by=" ORDER BY s.created_at DESC, i.item_index",
))

_register(ReportDef(
    key="scan_errors",
    name="Scan errors",
    description="All scan/batch errors with kind, code and acknowledgement state for error analysis.",
    scope="owner",
    columns=[
        ReportColumn("se.id", "Error ID"),
        ReportColumn("se.user_id", "Owner user ID", sensitive=True),
        ReportColumn("se.kind", "Kind"),
        ReportColumn("se.code", "Code"),
        ReportColumn("se.title", "Title"),
        ReportColumn("se.message", "Message"),
        ReportColumn("se.batch_id", "Batch ID"),
        ReportColumn("se.item_index", "Item index"),
        ReportColumn("se.receipt_id", "Receipt ID"),
        ReportColumn("se.read_at", "Read at"),
        ReportColumn("se.created_at", "Created"),
    ],
    base="scan_errors se",
    tenant="se.user_id",
    date_col="se.created_at",
    filters={"kind": "se.kind", "code": "se.code"},
    order_by=" ORDER BY se.created_at DESC",
))

_register(ReportDef(
    key="tasks",
    name="Batch tasks",
    description="Async batch processing log: task type, progress, errors and timing.",
    scope="owner",
    columns=[
        ReportColumn("t.id", "Task ID"),
        ReportColumn("t.user_id", "Owner user ID", sensitive=True),
        ReportColumn("t.task_type", "Task type"),
        ReportColumn("t.batch_title", "Batch title"),
        ReportColumn("t.status", "Status"),
        ReportColumn("t.total_items", "Total items"),
        ReportColumn("t.completed_items", "Completed items"),
        ReportColumn("t.percentage", "Percentage"),
        ReportColumn("t.start_timediff", "Placeholder", sensitive=True),
        ReportColumn("t.message", "Message"),
        ReportColumn("t.error", "Error", sensitive=True),
        ReportColumn("t.created_at", "Created"),
        ReportColumn("t.started_at", "Started"),
        ReportColumn("t.completed_at", "Completed"),
    ],
    base="tasks t",
    tenant="t.user_id",
    date_col="t.created_at",
    filters={"status": "t.status", "task_type": "t.task_type"},
    order_by=" ORDER BY t.created_at DESC",
))

_register(ReportDef(
    key="conversations",
    name="Conversation register",
    description="Messaging threads: participants (masked), linked receipt and activity.",
    scope="owner",
    columns=[
        ReportColumn("c.id", "Conversation ID"),
        ReportColumn("c.user_a", "Participant A", sensitive=True),
        ReportColumn("c.user_b", "Participant B", sensitive=True),
        ReportColumn("c.receipt_id", "Linked receipt"),
        ReportColumn("c.kind", "Kind"),
        ReportColumn("c.last_message_at", "Last message"),
        ReportColumn("c.created_at", "Created"),
    ],
    base="conversations c",
    tenant="(c.user_a = {uid} OR c.user_b = {uid})",
    filters={"kind": "c.kind"},
    order_by=" ORDER BY c.last_message_at DESC",
))

_register(ReportDef(
    key="messages",
    name="Message log",
    description="Every message sent through the platform with direction, kind and payload.",
    scope="admin",
    columns=[
        ReportColumn("m.id", "Message ID"),
        ReportColumn("m.conversation_id", "Conversation ID"),
        ReportColumn("m.sender_id", "Sender user ID", sensitive=True),
        ReportColumn("m.recipient_id", "Recipient user ID", sensitive=True),
        ReportColumn("m.kind", "Kind"),
        ReportColumn("m.body", "Body"),
        ReportColumn("m.payload", "Payload"),
        ReportColumn("m.read_at", "Read at"),
        ReportColumn("m.created_at", "Created"),
    ],
    base="messages m",
    date_col="m.created_at",
    filters={"kind": "m.kind"},
    order_by=" ORDER BY m.created_at DESC",
))

_register(ReportDef(
    key="audit_trail",
    name="Audit trail",
    description="Full audit log: every action across receipts with actor and change payloads.",
    scope="admin",
    columns=[
        ReportColumn("a.id", "Entry ID"),
        ReportColumn("a.timestamp", "Timestamp"),
        ReportColumn("a.action", "Action"),
        ReportColumn("a.changed_by", "Changed by user ID", sensitive=True),
        ReportColumn("a.user_id", "Receipt owner user ID", sensitive=True),
        ReportColumn("a.receipt_id", "Receipt ID"),
        ReportColumn("a.changes", "Changes"),
    ],
    base="audit_logs a",
    date_col="a.timestamp",
    filters={"action": "a.action"},
    order_by=" ORDER BY a.timestamp DESC",
))

_register(ReportDef(
    key="backups",
    name="Backup inventory",
    description="Data backup exports: files, sizes, image coverage and missing-image gaps.",
    scope="admin",
    columns=[
        ReportColumn("b.id", "Backup ID"),
        ReportColumn("b.user_id", "Owner user ID", sensitive=True),
        ReportColumn("b.filename", "Filename"),
        ReportColumn("b.size_bytes", "Size (bytes)"),
        ReportColumn("b.image_count", "Image count"),
        ReportColumn("b.missing_images", "Missing images"),
        ReportColumn("b.created_at", "Created"),
    ],
    base="backups b",
    date_col="b.created_at",
    order_by=" ORDER BY b.created_at DESC",
))

_register(ReportDef(
    key="review_batches",
    name="Review batches",
    description="Review queue uploads: CSV sources with per-batch item counts.",
    scope="owner",
    columns=[
        ReportColumn("rb.id", "Batch ID"),
        ReportColumn("rb.user_id", "Owner user ID", sensitive=True),
        ReportColumn("rb.name", "Name"),
        ReportColumn("rb.csv_filename", "CSV file"),
        ReportColumn("(SELECT COUNT(*) FROM review_batch_items x WHERE x.batch_id = rb.id)", "Item count"),
        ReportColumn("rb.created_at", "Created"),
        ReportColumn("rb.updated_at", "Updated"),
    ],
    base="review_batches rb",
    tenant="rb.user_id",
    date_col="rb.created_at",
    order_by=" ORDER BY rb.created_at DESC",
))

_register(ReportDef(
    key="review_items",
    name="Review item decisions",
    description="Per-item review outcomes and reviewer notes across all batches.",
    scope="owner",
    columns=[
        ReportColumn("ri.id", "Item ID"),
        ReportColumn("rb.id", "Batch ID"),
        ReportColumn("rb.name", "Batch name"),
        ReportColumn("rb.user_id", "Owner user ID", sensitive=True),
        ReportColumn("ri.receipt_id", "Receipt ID"),
        ReportColumn("ri.review_status", "Review status"),
        ReportColumn("ri.reviewer_notes", "Reviewer notes"),
        ReportColumn("ri.reviewed_at", "Reviewed at"),
    ],
    base="review_batch_items ri JOIN review_batches rb ON rb.id = ri.batch_id",
    tenant="rb.user_id",
    date_col="ri.reviewed_at",
    filters={"review_status": "ri.review_status"},
    order_by=" ORDER BY ri.reviewed_at DESC NULLS LAST, ri.id",
))

_register(ReportDef(
    key="locations",
    name="Location register",
    description="Configured locations/branches and their active state.",
    scope="owner",
    columns=[
        ReportColumn("l.id", "Location ID"),
        ReportColumn("l.name", "Name"),
        ReportColumn("l.is_active", "Active"),
        ReportColumn("l.created_by", "Created by user ID", sensitive=True),
        ReportColumn("l.created_at", "Created"),
        ReportColumn("l.updated_at", "Updated"),
    ],
    base="locations l",
    date_col="l.created_at",
    filters={"is_active": "l.is_active"},
    order_by=" ORDER BY l.name",
))

_register(ReportDef(
    key="users",
    name="User & staff accounts",
    description="Account register: admins vs regular users (identities masked by default).",
    scope="admin",
    columns=[
        ReportColumn("u.uid", "User ID", sensitive=True),
        ReportColumn("u.email", "Email", sensitive=True),
        ReportColumn("u.is_admin", "Is admin"),
        ReportColumn("u.display_name", "Display name", sensitive=True),
        ReportColumn("u.created_at", "Created"),
    ],
    base="users u",
    filters={"is_admin": "u.is_admin"},
    order_by=" ORDER BY u.created_at DESC",
))

_register(ReportDef(
    key="app_settings",
    name="Global configuration",
    description="Application-level settings key/value pairs (values masked by default).",
    scope="admin",
    columns=[
        ReportColumn("s.key", "Key"),
        ReportColumn("s.value", "Value", sensitive=True),
        ReportColumn("s.updated_at", "Updated"),
    ],
    base="app_settings s",
    date_col="s.updated_at",
    filters={"key": "s.key"},
    order_by=" ORDER BY s.key",
))

_register(ReportDef(
    key="user_configs",
    name="User configurations",
    description="Per-user preferences and AI provider settings (config values masked by default).",
    scope="admin",
    columns=[
        ReportColumn("c.user_id", "User ID", sensitive=True),
        ReportColumn("c.config_type", "Config type"),
        ReportColumn("c.provider", "AI provider"),
        ReportColumn("c.model_id", "AI model"),
        ReportColumn("c.default_tax_rate", "Default tax rate"),
        ReportColumn("c.configs", "Config values", sensitive=True),
        ReportColumn("c.updated_at", "Updated"),
    ],
    base=(
        "(SELECT up.user_id, 'preferences' AS config_type, NULL AS provider, "
        "NULL AS model_id, up.default_tax_rate, NULL AS configs, up.updated_at "
        "FROM user_preferences up "
        "UNION ALL "
        "SELECT ua.user_id, 'ai', ua.provider, ua.model_id, NULL, "
        "ua.configs::text, ua.updated_at FROM user_ai_settings ua) c"
    ),
    order_by=" ORDER BY c.updated_at DESC",
))

SCALAR_TYPES = (datetime, date, Decimal, UUID)


def _serialize(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, (dict, list)) and not isinstance(value, (str, bytes)):
        try:
            return json.dumps(value, default=str)
        except Exception:
            return str(value)
    return value


def _coerce_bool(v: str) -> str:
    if v.lower() not in ("true", "false"):
        raise ValueError("boolean filters accept 'true' or 'false'")
    return v.lower()


def _parse_date(v: str) -> date:
    if not v:
        return None
    try:
        return datetime.strptime(v, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("dates must be in YYYY-MM-DD format")


def _alias(col: str) -> str:
    """Stable, unique output-column key derived from the SQL fragment.

    Postgres strips table qualifiers from plain SELECT lists, so every
    selected column is aliased explicitly to keep row keys deterministic
    across reports and formats.
    """
    return re.sub(r"\W", "_", col).strip("_")


async def list_report_defs(is_admin: bool) -> List[Dict[str, Any]]:
    """Catalog visible to the requesting user (owner reports for everyone, admin ones only for admins)."""
    out = []
    for key in sorted(REPORT_DEFS):
        d = REPORT_DEFS[key]
        if d.scope == "admin" and not is_admin:
            continue
        out.append({
            "key": d.key,
            "name": d.name,
            "description": d.description,
            "scope": d.scope,
            "columns": [
                {"key": _alias(c.col), "label": c.label, "sensitive": c.sensitive}
                for c in d.columns
            ],
            "filters": sorted(d.filters),
            "dateFilter": d.date_col is not None,
        })
    return out


async def run_report(
    current_uid: str,
    is_admin: bool,
    key: str,
    report_format: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    include_sensitive: bool = False,
    filters: Optional[Dict[str, str]] = None,
) -> Dict[str, Any]:
    """Run a report definition with tenant scoping, masking and format rendering."""
    d = REPORT_DEFS.get(key)
    if d is None:
        raise LookupError(f"unknown report: {key}")
    if d.scope == "admin" and not is_admin:
        raise PermissionError("report requires admin access")
    if include_sensitive and not is_admin:
        raise PermissionError("include_sensitive requires admin access")

    columns = [c for c in d.columns if not c.sensitive or include_sensitive]
    if not columns:
        raise ValueError("report has no exportable columns")

    where: List[str] = []
    args: List[Any] = []

    if not is_admin and d.tenant:
        if "{uid}" in d.tenant:
            where.append(d.tenant.replace("{uid}", "$" + str(len(args) + 1)))
            args.append(current_uid)
        else:
            where.append(f"{d.tenant} = $" + str(len(args) + 1))
            args.append(current_uid)

    date_prefix = "(" + d.date_col + ")" if d.date_col else None
    if date_prefix:
        if date_from:
            where.append(f"{date_prefix} >= $" + str(len(args) + 1))
            args.append(_parse_date(date_from))
        if date_to:
            where.append(
                f"{date_prefix} < ($" + str(len(args) + 1) + "::timestamp + interval '1 day')"
            )
            args.append(_parse_date(date_to))

    if filters:
        for name, value in filters.items():
            col = d.filters.get(name)
            if col is None:
                raise ValueError(f"unknown filter: {name}")
            if len(value) > 200:
                raise ValueError(f"filter value too long: {name}")
            if name == "is_active" or name == "is_admin":
                value = _coerce_bool(value)
            where.append(f"{col} = $" + str(len(args) + 1))
            args.append(value)

    if d.expense_cond:
        where.append(d.expense_cond)

    select = ", ".join(f"{c.col} AS {_alias(c.col)}" for c in columns)
    sql = f"SELECT {select} FROM {d.base}"
    if where:
        sql += " WHERE " + " AND ".join(where)
    if d.group_by:
        grouped = {part.strip() for part in d.group_by.split(",")}
        selected_fragments = {c.col for c in columns}
        if not grouped <= selected_fragments:
            raise ValueError("grouping columns were masked out of this report")
        sql += f" GROUP BY {d.group_by}"
    sql += d.order_by + f" LIMIT {MAX_ROWS}"

    pool = await get_pool()
    async with pool.acquire() as conn:
        raw_rows = await conn.fetch(sql, *args)

    rows = [{k: _serialize(v) for k, v in dict(r).items()} for r in raw_rows]
    truncated = len(rows) >= MAX_ROWS

    summary: Dict[str, Any] = {"row_count": len(rows), "truncated": truncated}
    if d.money_cols and rows and not truncated:
        col_by_label = {c.label: _alias(c.col) for c in d.columns}
        for label in d.money_cols:
            col_key = col_by_label.get(label, label)
            numeric = [
                float(row[col_key])
                for row in rows
                if row.get(col_key) is not None and isinstance(row.get(col_key), (int, float))
            ]
            if numeric:
                summary[label] = round(sum(numeric), 2)

    try:
        await AuditService.log(
            user_id=current_uid,
            receipt_id=None,
            action=AuditAction.REPORT_EXPORT,
            changed_by=current_uid,
            changes=[
                AuditFieldChange(
                    field="report",
                    old_value=None,
                    new_value=f"{key} ({report_format})",
                )
            ],
        )
    except Exception as e:  # never let auditing break an export
        logger.warning("report export audit failed for %s: %s", key, e)

    return {
        "report": d.key,
        "name": d.name,
        "format": report_format,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "columns": [{"key": _alias(c.col), "label": c.label} for c in columns],
        "rows": rows,
        "summary": summary,
    }


# ── Format renderers (csv / xlsx / pdf / json) ───────────────────────────────


def render_csv(payload: Dict[str, Any]) -> bytes:
    headers = [c["label"] for c in payload["columns"]]
    buf = io.StringIO()
    buf.write("\ufeff")
    writer = csv.writer(buf)
    writer.writerow(headers)
    for row in payload["rows"]:
        writer.writerow(["" if v is None else v for v in row.values()])
    if payload["summary"].get("row_count") and not payload["summary"].get("truncated"):
        writer.writerow([])
        writer.writerow(["TOTAL ROWS", payload["summary"]["row_count"]])
        for label in payload["summary"]:
            if label in ("row_count", "truncated"):
                continue
            writer.writerow([f"Total {label}", payload["summary"][label]])
    return buf.getvalue().encode("utf-8")


def render_xlsx(payload: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = payload["report"][:31] or "report"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    headers = [c["label"] for c in payload["columns"]]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for row in payload["rows"]:
        ws.append(["" if v is None else v for v in row.values()])
    if payload["summary"].get("row_count") and not payload["summary"].get("truncated"):
        ws.append([])
        ws.append(["TOTAL ROWS", payload["summary"]["row_count"]])
        for label in payload["summary"]:
            if label in ("row_count", "truncated"):
                continue
            ws.append([f"Total {label}", payload["summary"][label]])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    for idx, header in enumerate(headers, start=1):
        widths = [
            len(str(row.get(header_column["key"], "")))
            for row in payload["rows"][:200]
        ]
        width = min(60, max([len(header)] + widths))
        ws.column_dimensions[get_column_letter(idx)].width = width + 2
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def render_pdf(payload: Dict[str, Any]) -> bytes:
    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=payload["name"],
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ReportTitle", parent=styles["Title"], fontSize=12)
    cell_style = ParagraphStyle(
        "Cell", parent=styles["BodyText"], fontSize=6, leading=8
    )
    head_style = ParagraphStyle(
        "Head", parent=cell_style, textColor=colors.white, fontName="Helvetica-Bold"
    )
    headers = [c["label"] for c in payload["columns"]]
    data = [[Paragraph(h, head_style) for h in headers]]
    for row in payload["rows"]:
        data.append([Paragraph(str("" if v is None else v), cell_style) for v in row.values()])
    story = [
        Paragraph(f"{payload['name']} — generated {payload['generated_at']}", title_style),
        Paragraph(f"Rows: {payload['summary'].get('row_count', 0)}", styles["BodyText"]),
    ]
    if payload["summary"].get("row_count") and not payload["summary"].get("truncated"):
        totals = ["TOTAL ROWS", str(payload["summary"]["row_count"])]
        for label in payload["summary"]:
            if label in ("row_count", "truncated"):
                continue
            totals.append(f"{label}: {payload['summary'][label]}")
        padded = totals + [""] * (len(headers) - len(totals))
        data.append([Paragraph(str(cell), cell_style) for cell in padded])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF3FA")]),
    ]))
    story.append(table)
    doc.build(story)
    return out.getvalue()


def render_json(payload: Dict[str, Any]) -> bytes:
    return json.dumps(payload, indent=2, default=str).encode("utf-8")


RENDERERS = {
    "csv": render_csv,
    "xlsx": render_xlsx,
    "pdf": render_pdf,
    "json": render_json,
}

CONTENT_TYPES = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
    "json": "application/json",
}