import csv
import io
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

logger = logging.getLogger(__name__)

RECEIPT_FIELDS = ["supplier", "totalAmount", "taxAmount", "receiptDate", "category", "invoiceNumber", "kraPin", "buyerKraPin", "cuInvoice", "batchTitle", "status"]
RECEIPT_LABELS = {
    "supplier": "Supplier", "totalAmount": "Total", "taxAmount": "Tax",
    "receiptDate": "Date", "category": "Category", "invoiceNumber": "Invoice #",
    "kraPin": "Seller PIN", "buyerKraPin": "Buyer PIN", "cuInvoice": "CU Invoice",
    "batchTitle": "Batch", "status": "Status",
}
MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def receipts_rows(receipts: List[dict], columns: Optional[List[str]] = None) -> List[dict]:
    """One row per receipt with selected fields."""
    cols = columns or RECEIPT_FIELDS
    rows = []
    for r in receipts:
        row = {}
        for c in cols:
            val = r.get(c, "")
            if c == "totalAmount":
                row[c] = sanitize_numeric(val)
            else:
                row[c] = val
        rows.append(row)
    return rows

# ─── Helpers ──────────────────────────────────────────────────────────────

def sanitize_numeric(v: Any) -> float:
    try:
        return float(str(v).replace(",", "").replace("KES", "").strip())
    except (ValueError, AttributeError):
        return 0.0


def parse_date_mdy(date_str: str) -> Optional[datetime]:
    """Parse MM/DD/YYYY string to datetime. Returns None if unparseable."""
    try:
        parts = str(date_str).split("/")
        if len(parts) == 3:
            m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
            return datetime(y, m, d)
    except (ValueError, AttributeError):
        pass
    return None


def extract_year(date_str: str) -> Optional[str]:
    parts = date_str.split("/")
    if len(parts) == 3:
        return parts[2]
    return None


def month_only(date_str: str) -> str:
    parts = date_str.split("/")
    if len(parts) == 3:
        m = int(parts[0])
        return MONTH_NAMES[m - 1] if 1 <= m <= 12 else date_str
    return date_str


def month_sort_key(date_str: str) -> str:
    parts = date_str.split("/")
    if len(parts) == 3:
        return f"{parts[2]}-{parts[0].zfill(2)}"
    return date_str or ""


def field_value(r: dict, field: str) -> str:
    if field == "month":
        return month_only(r.get("receiptDate", ""))
    if field == "supplier":
        return r.get("supplier", "Unknown")
    if field == "category":
        return r.get("category", "Other")
    return "Unknown"


def years_in_data(receipts: List[dict]) -> List[str]:
    years = set()
    for r in receipts:
        y = extract_year(r.get("receiptDate", ""))
        if y:
            years.add(y)
    return sorted(years)


# ─── Aggregation ──────────────────────────────────────────────────────────

def _normalize_receipt_date(r: dict) -> str:
    """Convert a receipt's MM/DD/YYYY date to YYYYMMDD for string comparison."""
    d = r.get("receiptDate", "") or ""
    parts = d.split("/")
    if len(parts) == 3:
        return parts[2] + parts[0].zfill(2) + parts[1].zfill(2)
    return d


def _normalize_filter_date(d: Optional[str]) -> Optional[str]:
    """Normalize a filter date (YYYY-MM-DD or MM/DD/YYYY) to YYYYMMDD."""
    if not d:
        return None
    if "-" in d:
        parts = d.split("-")
        if len(parts) == 3:
            return parts[0] + parts[1].zfill(2) + parts[2].zfill(2)
    if "/" in d:
        parts = d.split("/")
        if len(parts) == 3:
            return parts[2] + parts[0].zfill(2) + parts[1].zfill(2)
    return d


def aggregate_receipts(
    receipts: List[dict],
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    category: Optional[str] = None,
    year: Optional[str] = None,
) -> List[dict]:
    n_from = _normalize_filter_date(date_from)
    n_to = _normalize_filter_date(date_to)
    filtered = receipts
    if n_from:
        filtered = [r for r in filtered if _normalize_receipt_date(r) >= n_from]
    if n_to:
        filtered = [r for r in filtered if _normalize_receipt_date(r) <= n_to]
    if category:
        filtered = [r for r in filtered if r.get("category") == category]
    if year:
        filtered = [r for r in filtered if extract_year(r.get("receiptDate", "")) == year]
    return filtered


def compute_pivot(
    receipts: List[dict],
    row_field: str,
    col_field: str,
    value_field: str,
) -> Tuple[List[str], List[str], Dict[str, Dict[str, float]], Dict[str, float], Dict[str, float], float]:
    grid: Dict[str, Dict[str, float]] = {}
    row_totals: Dict[str, float] = {}
    col_totals: Dict[str, float] = {}
    grand_total = 0.0
    row_set: set = set()
    col_set: set = set()

    for r in receipts:
        rv = field_value(r, row_field)
        cv = field_value(r, col_field)
        row_set.add(rv)
        col_set.add(cv)
        val = 1.0 if value_field == "count" else sanitize_numeric(r.get("totalAmount"))
        if rv not in grid:
            grid[rv] = {}
        grid[rv][cv] = grid[rv].get(cv, 0) + val
        row_totals[rv] = row_totals.get(rv, 0) + val
        col_totals[cv] = col_totals.get(cv, 0) + val
        grand_total += val

    row_labels = sorted(row_set, key=lambda x: -row_totals.get(x, 0))
    col_labels = sorted(col_set, key=lambda x: -col_totals.get(x, 0))
    return row_labels, col_labels, grid, row_totals, col_totals, grand_total


def category_breakdown(receipts: List[dict]) -> List[dict]:
    cats: Dict[str, float] = {}
    cnts: Dict[str, int] = {}
    for r in receipts:
        c = r.get("category", "Other")
        cats[c] = cats.get(c, 0) + sanitize_numeric(r.get("totalAmount"))
        cnts[c] = cnts.get(c, 0) + 1
    total = sum(cats.values())
    return [
        {"Category": c, "Total": round(v, 2), "Count": cnts[c], "Pct": round(v / total * 100, 1) if total else 0}
        for c, v in sorted(cats.items(), key=lambda x: -x[1])
    ]


def supplier_breakdown(receipts: List[dict]) -> List[dict]:
    sups: Dict[str, float] = {}
    cnts: Dict[str, int] = {}
    for r in receipts:
        s = r.get("supplier", "Unknown")
        sups[s] = sups.get(s, 0) + sanitize_numeric(r.get("totalAmount"))
        cnts[s] = cnts.get(s, 0) + 1
    return [
        {"Supplier": s, "Total": round(v, 2), "Count": cnts[s], "Avg": round(v / cnts[s], 2)}
        for s, v in sorted(sups.items(), key=lambda x: -x[1])
    ]


def monthly_rows(receipts: List[dict]) -> List[dict]:
    months: Dict[str, float] = {}
    cnts: Dict[str, int] = {}
    for r in receipts:
        m = month_only(r.get("receiptDate", ""))
        sk = month_sort_key(r.get("receiptDate", ""))
        months[sk] = months.get(sk, 0) + sanitize_numeric(r.get("totalAmount"))
        cnts[sk] = cnts.get(sk, 0) + 1
    return [
        {"Month": month_only(sk[5:] if "-" in sk else sk), "Total": round(v, 2), "Count": cnts[sk], "Avg": round(v / cnts[sk], 2)}
        for sk, v in sorted(months.items())
    ]


def detailed_rows(receipts: List[dict]) -> List[dict]:
    rows = []
    for r in receipts:
        items = r.get("items") or []
        base = {
            "Receipt ID": r.get("id", ""),
            "Date": r.get("receiptDate", ""),
            "Supplier": r.get("supplier", ""),
            "Category": r.get("category", ""),
            "Invoice": r.get("invoiceNumber", ""),
            "Receipt Total": sanitize_numeric(r.get("totalAmount")),
        }
        if items:
            for item in items:
                qty = sanitize_numeric(item.get("quantity", 1)) or 1
                price = sanitize_numeric(item.get("price"))
                tax = sanitize_numeric(item.get("tax"))
                discount_pct = sanitize_numeric(item.get("discount"))
                subtotal = qty * (price + tax)
                discount_factor = (1 - discount_pct / 100) if discount_pct else 1
                row = dict(base)
                row["Item"] = item.get("name", "")
                row["Qty"] = qty
                row["Price"] = price
                row["Tax"] = tax
                row["Disc%"] = discount_pct if discount_pct else ""
                row["Item Total"] = round(subtotal * discount_factor, 2)
                rows.append(row)
        else:
            row = dict(base)
            row["Item"] = ""
            row["Qty"] = 0
            row["Price"] = 0
            row["Tax"] = 0
            row["Disc%"] = ""
            row["Item Total"] = 0
            rows.append(row)
    return rows


def tax_summary(receipts: List[dict]) -> List[dict]:
    zero_amt = 0.0
    taxed_amt = 0.0
    zero_cnt = 0
    taxed_cnt = 0
    for r in receipts:
        for item in r.get("items") or []:
            qty = sanitize_numeric(item.get("quantity", 1)) or 1
            price = sanitize_numeric(item.get("price"))
            tax = sanitize_numeric(item.get("tax"))
            total = qty * (price + tax)
            if item.get("isZeroRated"):
                zero_amt += total
                zero_cnt += 1
            else:
                taxed_amt += total
                taxed_cnt += 1
    return [
        {"Category": "Zero Rated (Tax Exempt)", "Total": round(zero_amt, 2), "Count": zero_cnt},
        {"Category": "Taxable", "Total": round(taxed_amt, 2), "Count": taxed_cnt},
        {"Category": "Combined Total", "Total": round(zero_amt + taxed_amt, 2), "Count": zero_cnt + taxed_cnt},
    ]


# ─── Helpers: does report involve month dimension? ───────────────────────

def _report_uses_month(report_type: str, pivot_config: Optional[dict] = None) -> bool:
    if report_type == "monthly":
        return True
    if report_type == "pivot" and pivot_config:
        return pivot_config.get("rowField") == "month" or pivot_config.get("colField") == "month"
    return False


# ─── Styles ───────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)


def style_header_row(ws, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def style_data_rows(ws, row_count: int, col_count: int):
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row in range(2, row_count + 1):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="right" if col > 1 else "left")
            if row % 2 == 0:
                cell.fill = alt_fill


def auto_width(ws, col_count: int, max_width: int = 40):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                v = cell.value
                if v:
                    if isinstance(v, datetime):
                        max_len = max(max_len, 10)  # MM/DD/YYYY
                    else:
                        max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, max_width)


# ═══════════════════════════════════════════════════════════════════════════
# Excel Generator
# ═══════════════════════════════════════════════════════════════════════════

def _build_pivot_rows(filtered, pc):
    row_labels, col_labels, grid, row_tots, col_tots, grand = compute_pivot(
        filtered, pc["rowField"], pc["colField"], pc["valueField"]
    )
    header = [pc["rowField"].capitalize()] + col_labels + ["Total"]
    data_rows = []
    for rl in row_labels:
        data_rows.append([rl] + [round(grid.get(rl, {}).get(cl, 0), 2) for cl in col_labels] + [round(row_tots.get(rl, 0), 2)])
    data_rows.append(["Grand Total"] + [round(col_tots.get(cl, 0), 2) for cl in col_labels] + [round(grand, 2)])
    return header, data_rows


def generate_excel(receipts: List[dict], report_type: str, date_from: Optional[str] = None,
                   date_to: Optional[str] = None, pivot_config: Optional[dict] = None,
                   columns: Optional[List[str]] = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    filtered = aggregate_receipts(receipts, date_from, date_to)

    if report_type == "receipts":
        cols = columns or RECEIPT_FIELDS
        headers = [RECEIPT_LABELS.get(c, c) for c in cols]
        data = []
        for r in receipts_rows(filtered, cols):
            data.append([r.get(c, "") for c in cols])
        _add_excel_sheet(wb, "Receipts", headers, data)

    if report_type == "detailed":
        _detail_headers = ["Receipt ID", "Date", "Supplier", "Category", "Invoice", "Item", "Qty", "Price", "Tax", "Disc%", "Item Total", "Receipt Total"]
        _detail_data = []
        for r in detailed_rows(filtered):
            row = []
            for h in _detail_headers:
                v = r.get(h, "")
                if h == "Date" and v:
                    parsed = parse_date_mdy(str(v))
                    row.append(parsed if parsed else v)
                else:
                    row.append(v)
            _detail_data.append(row)
        _add_excel_sheet(wb, "Detailed", _detail_headers, _detail_data, date_cols=[1])
        _add_excel_sheet(wb, "By Category", ["Category", "Total", "Count", "Percentage"],
                         [[r["Category"], r["Total"], r["Count"], r["Pct"]] for r in category_breakdown(filtered)])
        _add_excel_sheet(wb, "By Supplier", ["Supplier", "Total", "Count", "Avg per Receipt"],
                         [[r["Supplier"], r["Total"], r["Count"], r["Avg"]] for r in supplier_breakdown(filtered)])
        # Monthly — one sheet per year
        for year in years_in_data(filtered):
            yf = aggregate_receipts(filtered, year=year)
            _add_excel_sheet(wb, f"Monthly {year}", ["Month", "Total", "Count", "Avg per Receipt"],
                             [[r["Month"], r["Total"], r["Count"], r["Avg"]] for r in monthly_rows(yf)])
        _add_excel_sheet(wb, "Tax Summary", ["Category", "Total", "Count"],
                         [[r["Category"], r["Total"], r["Count"]] for r in tax_summary(filtered)])

    elif report_type == "pivot" and pivot_config:
        pc = pivot_config
        if pc.get("rowField") == "month" or pc.get("colField") == "month":
            for year in years_in_data(filtered):
                yf = aggregate_receipts(filtered, year=year)
                header, data_rows = _build_pivot_rows(yf, pc)
                _add_excel_sheet(wb, f"Pivot {year}", header, data_rows)
        else:
            header, data_rows = _build_pivot_rows(filtered, pc)
            _add_excel_sheet(wb, "Pivot", header, data_rows)

    elif report_type == "monthly":
        for year in years_in_data(filtered):
            yf = aggregate_receipts(filtered, year=year)
            _add_excel_sheet(wb, f"Monthly {year}", ["Month", "Total", "Count", "Avg per Receipt"],
                             [[r["Month"], r["Total"], r["Count"], r["Avg"]] for r in monthly_rows(yf)])

    else:
        mapper = {
            "category": ("By Category", ["Category", "Total", "Count", "Percentage"],
                         [[r["Category"], r["Total"], r["Count"], r["Pct"]] for r in category_breakdown(filtered)]),
            "supplier": ("By Supplier", ["Supplier", "Total", "Count", "Avg"],
                         [[r["Supplier"], r["Total"], r["Count"], r["Avg"]] for r in supplier_breakdown(filtered)]),
            "tax": ("Tax Summary", ["Category", "Total", "Count"],
                    [[r["Category"], r["Total"], r["Count"]] for r in tax_summary(filtered)]),
        }
        name, h, d = mapper.get(report_type, ("Report", ["Field", "Value"], [["No data", ""]]))
        _add_excel_sheet(wb, name, h, d)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _add_excel_sheet(wb: Workbook, name: str, headers: List[str], data: List[list], date_cols: Optional[List[int]] = None):
    ws = wb.create_sheet(title=name[:31])
    ws.append(headers)
    for row in data:
        ws.append(row)
    style_header_row(ws, len(headers))
    style_data_rows(ws, len(data) + 1, len(headers))
    if date_cols:
        for col_idx in date_cols:
            for row in range(2, len(data) + 2):
                cell = ws.cell(row=row, column=col_idx + 1)
                if isinstance(cell.value, datetime):
                    cell.number_format = 'MM/DD/YYYY'
    auto_width(ws, len(headers))


# ═══════════════════════════════════════════════════════════════════════════
# PDF Generator
# ═══════════════════════════════════════════════════════════════════════════

def _make_table(doc, headers: List[str], rows: List[list], cell_style) -> Table:
    p_headers = [Paragraph(h, cell_style) for h in headers]
    p_rows = [[Paragraph(str(c), cell_style) for c in row] for row in rows]
    table_data = [p_headers] + p_rows
    col_w = (doc.width - 20 * mm) / len(headers)
    table = Table(table_data, colWidths=[col_w] * len(headers), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ]))
    return table


def generate_pdf(receipts: List[dict], report_type: str, date_from: Optional[str] = None,
                 date_to: Optional[str] = None, pivot_config: Optional[dict] = None,
                 columns: Optional[List[str]] = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=15*mm, bottomMargin=20*mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Heading1"], fontSize=16, textColor=colors.HexColor("#1E40AF"), spaceAfter=6)
    meta_style = ParagraphStyle("Meta", parent=styles["Normal"], fontSize=8, textColor=colors.gray, spaceAfter=4)
    section_style = ParagraphStyle("Section", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#1E40AF"), spaceAfter=4, spaceBefore=8)
    cell_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)

    elements = []
    filtered = aggregate_receipts(receipts, date_from, date_to)
    total_spent = sum(sanitize_numeric(r.get("totalAmount")) for r in filtered)

    title_map = {
        "detailed": "Detailed Receipt Report", "category": "Spending by Category",
        "supplier": "Spending by Supplier", "monthly": "Monthly Spending Trend",
        "tax": "Tax Summary Report", "pivot": "Pivot Table",
    }
    title = title_map.get(report_type, "Report")
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | {len(filtered)} receipts | Total: {total_spent:,.2f}", meta_style))
    if date_from or date_to:
        elements.append(Paragraph(f"Period: {date_from or '…'} — {date_to or '…'}", meta_style))

    uses_month = _report_uses_month(report_type, pivot_config)

    if uses_month:
        # One section per year
        for year in years_in_data(filtered):
            yf = aggregate_receipts(filtered, year=year)
            if not yf:
                continue
            elements.append(Spacer(1, 4*mm))
            if len(elements) > 3:
                elements.append(PageBreak())
            elements.append(Paragraph(f"{year}", section_style))

            if report_type == "monthly":
                data = monthly_rows(yf)
                headers = ["Month", "Total", "Count", "Avg"]
                rows = [[r["Month"], f"{r['Total']:,.2f}", r["Count"], f"{r['Avg']:,.2f}"] for r in data]
            elif report_type == "pivot" and pivot_config:
                pc = pivot_config
                rl, cl, grid, rt, ct, gt = compute_pivot(yf, pc["rowField"], pc["colField"], pc["valueField"])
                headers = [pc["rowField"].capitalize()] + cl + ["Total"]
                rows = []
                for rv in rl:
                    rows.append([rv] + [f"{grid.get(rv, {}).get(cv, 0):,.2f}" for cv in cl] + [f"{rt.get(rv, 0):,.2f}"])
                rows.append(["Grand Total"] + [f"{ct.get(cv, 0):,.2f}" for cv in cl] + [f"{gt:,.2f}"])
            else:
                continue

            year_total = sum(sanitize_numeric(r.get("totalAmount")) for r in yf)
            elements.append(Paragraph(f"Year Total: {year_total:,.2f} | Receipts: {len(yf)}", meta_style))
            if rows:
                elements.append(_make_table(doc, headers, rows, cell_style))

    elif report_type == "detailed":
        elements.append(Spacer(1, 4*mm))
        headers = ["Receipt ID", "Date", "Supplier", "Category", "Invoice", "Item", "Qty", "Price", "Tax", "Disc%", "Item Total"]
        rows = [[r.get(h, "") for h in headers]
                for r in detailed_rows(filtered)]
        if rows:
            elements.append(_make_table(doc, headers, rows, cell_style))

    elif report_type == "receipts":
        elements.append(Spacer(1, 4*mm))
        cols = columns or RECEIPT_FIELDS
        headers = [RECEIPT_LABELS.get(c, c) for c in cols]
        rows = [[str(r.get(c, "")) for c in cols] for r in receipts_rows(filtered, cols)]
        if rows:
            elements.append(_make_table(doc, headers, rows, cell_style))

    elif report_type == "category":
        elements.append(Spacer(1, 4*mm))
        data = category_breakdown(filtered)
        headers = ["Category", "Total", "Count", "Percentage"]
        rows = [[r["Category"], f"{r['Total']:,.2f}", r["Count"], f"{r['Pct']}%"] for r in data]
        if rows:
            elements.append(_make_table(doc, headers, rows, cell_style))

    elif report_type == "supplier":
        elements.append(Spacer(1, 4*mm))
        data = supplier_breakdown(filtered)
        headers = ["Supplier", "Total", "Count", "Avg"]
        rows = [[r["Supplier"], f"{r['Total']:,.2f}", r["Count"], f"{r['Avg']:,.2f}"] for r in data]
        if rows:
            elements.append(_make_table(doc, headers, rows, cell_style))

    elif report_type == "tax":
        elements.append(Spacer(1, 4*mm))
        data = tax_summary(filtered)
        headers = ["Category", "Total", "Count"]
        rows = [[r["Category"], f"{r['Total']:,.2f}", r["Count"]] for r in data]
        if rows:
            elements.append(_make_table(doc, headers, rows, cell_style))

    doc.build(elements)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════════════
# CSV Generator
# ═══════════════════════════════════════════════════════════════════════════

def generate_csv(receipts: List[dict], report_type: str, date_from: Optional[str] = None,
                 date_to: Optional[str] = None, pivot_config: Optional[dict] = None,
                 columns: Optional[List[str]] = None) -> bytes:
    filtered = aggregate_receipts(receipts, date_from, date_to)
    buf = io.StringIO()
    writer = csv.writer(buf)

    uses_month = _report_uses_month(report_type, pivot_config)

    if report_type == "detailed":
        _detail_headers = ["Receipt ID", "Date", "Supplier", "Category", "Invoice", "Item", "Qty", "Price", "Tax", "Disc%", "Item Total", "Receipt Total"]
        writer.writerow(_detail_headers)
        for r in detailed_rows(filtered):
            writer.writerow([r.get(h, "") for h in _detail_headers])

    elif report_type == "receipts":
        cols = columns or RECEIPT_FIELDS
        headers = [RECEIPT_LABELS.get(c, c) for c in cols]
        writer.writerow(headers)
        for r in receipts_rows(filtered, cols):
            writer.writerow([r.get(c, "") for c in cols])

    elif report_type == "pivot" and pivot_config:
        pc = pivot_config
        if uses_month:
            for year in years_in_data(filtered):
                yf = aggregate_receipts(filtered, year=year)
                rl, cl, grid, rt, ct, gt = compute_pivot(yf, pc["rowField"], pc["colField"], pc["valueField"])
                writer.writerow([f"--- {year} ---"])
                writer.writerow([pc["rowField"].capitalize()] + cl + ["Total"])
                for rv in rl:
                    writer.writerow([rv] + [round(grid.get(rv, {}).get(cv, 0), 2) for cv in cl] + [round(rt.get(rv, 0), 2)])
                writer.writerow(["Grand Total"] + [round(ct.get(cv, 0), 2) for cv in cl] + [round(gt, 2)])
                writer.writerow([])
        else:
            rl, cl, grid, rt, ct, gt = compute_pivot(filtered, pc["rowField"], pc["colField"], pc["valueField"])
            writer.writerow([pc["rowField"].capitalize()] + cl + ["Total"])
            for rv in rl:
                writer.writerow([rv] + [round(grid.get(rv, {}).get(cv, 0), 2) for cv in cl] + [round(rt.get(rv, 0), 2)])
            writer.writerow(["Grand Total"] + [round(ct.get(cv, 0), 2) for cv in cl] + [round(gt, 2)])

    elif report_type == "monthly":
        writer.writerow(["Year", "Month", "Total", "Count", "Avg per Receipt"])
        for year in years_in_data(filtered):
            yf = aggregate_receipts(filtered, year=year)
            for r in monthly_rows(yf):
                writer.writerow([year, r["Month"], r["Total"], r["Count"], r["Avg"]])

    elif report_type == "category":
        writer.writerow(["Category", "Total", "Count", "Percentage"])
        for r in category_breakdown(filtered):
            writer.writerow([r["Category"], r["Total"], r["Count"], r["Pct"]])
    elif report_type == "supplier":
        writer.writerow(["Supplier", "Total", "Count", "Avg per Receipt"])
        for r in supplier_breakdown(filtered):
            writer.writerow([r["Supplier"], r["Total"], r["Count"], r["Avg"]])
    elif report_type == "tax":
        writer.writerow(["Category", "Total", "Count"])
        for r in tax_summary(filtered):
            writer.writerow([r["Category"], r["Total"], r["Count"]])

    return buf.getvalue().encode("utf-8")


# ═══════════════════════════════════════════════════════════════════════════
# Dispatcher
# ═══════════════════════════════════════════════════════════════════════════

def generate_export(
    receipts: List[dict],
    export_format: str,
    report_type: str,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    pivot_config: Optional[dict] = None,
    columns: Optional[List[str]] = None,
) -> bytes:
    generators = {
        "xlsx": generate_excel,
        "pdf": generate_pdf,
        "csv": generate_csv,
    }
    gen = generators.get(export_format)
    if not gen:
        raise ValueError(f"Unsupported export format: {export_format}")
    return gen(receipts, report_type, date_from, date_to, pivot_config, columns)
